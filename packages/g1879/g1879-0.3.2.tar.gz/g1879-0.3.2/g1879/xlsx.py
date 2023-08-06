#!/usr/bin/env python
# -*- coding:utf-8 -*-
from copy import copy
from pathlib import Path
from typing import Union

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from win32com import client

from .paths import make_valid_name


def xls_to_xlsx(path: str, del_src: bool = True) -> bool:
    """把xls格式另存为xlsx格式
    :param path: xls文件路径
    :param del_src: 另存之后是否删除源文件
    :return: 是否成功
    """
    excel = None

    try:
        excel = client.Dispatch("Ket.Application")  # 调用WPS程序
    except:
        try:
            excel = client.gencache.EnsureDispatch('Excel.Application')  # 调用Office程序
        except:
            pass

    if excel:
        wb = excel.Workbooks.Open(path)
        wb.SaveAs(f"{path}x", 51)
        wb.Close()

        if del_src:
            Path(path).unlink()

        return True

    return False


def copy_cell(src_cell: Cell,
              tar_sheet: Worksheet,
              row_or_loc: Union[str, int],
              col: Union[str, int] = None) -> None:
    """复制单元格到目标工作表
    :param src_cell: 源单元格
    :param tar_sheet: 目标工作表
    :param row_or_loc: 目标工作表中单元格行号或位置字符串
    :param col: 目标工作表中单元格位置
    :return: None
    """
    if isinstance(row_or_loc, str):
        tar_cell = tar_sheet[row_or_loc]
    elif isinstance(row_or_loc, int) and isinstance(col, int):
        tar_cell = tar_sheet.cell(row_or_loc, col)
    elif isinstance(row_or_loc, int) and isinstance(col, str):
        tar_cell = tar_sheet[f'{col}{row_or_loc}']
    else:
        raise ValueError('传入的位置参数不正确')

    tar_cell.value = src_cell.value
    tar_cell.hyperlink = src_cell.hyperlink

    if src_cell.has_style:
        tar_cell.font = copy(src_cell.font)
        tar_cell.border = copy(src_cell.border)
        tar_cell.fill = copy(src_cell.fill)
        tar_cell.number_format = copy(src_cell.number_format)
        tar_cell.protection = copy(src_cell.protection)
        tar_cell.alignment = copy(src_cell.alignment)


def copy_row(src_sheet: Worksheet,
             row_or_num: Union[int, tuple],
             tar_sheet: Worksheet,
             tar_row_num: int = None) -> None:
    """复制一行到目标工作表
    :param src_sheet: 源工作表
    :param row_or_num: 源行对象或行号
    :param tar_sheet: 目标工作表
    :param tar_row_num: 目标行号，默认为在最后添加一行
    :return: None
    """
    if not tar_row_num:
        max_row = tar_sheet.max_row
        if len(tar_sheet[max_row]) == 1 and tar_sheet[max_row][0].value is None:
            tar_row_num = 1
        else:
            tar_row_num = max_row + 1

    tar_row_num = tar_row_num or tar_sheet.max_row + 1
    if isinstance(row_or_num, tuple):
        row = row_or_num
        num = row[0].row
    else:
        row = src_sheet[row_or_num]
        num = row_or_num

    tar_sheet.row_dimensions[tar_row_num].height = src_sheet.row_dimensions[num].height

    for key, cell in enumerate(row, 1):
        col = get_column_letter(key)
        tar_sheet.column_dimensions[col].width = src_sheet.column_dimensions[col].width
        copy_cell(cell, tar_sheet, tar_row_num, key)


def split_sheet(sheet: Worksheet,
                col: Union[str, int],
                save_path: Union[str, Path],
                begin_row: int = 2,
                keys: Union[str, tuple, list, set] = None) -> None:
    """按照某列内容把xlsx文件拆分成多个文件
    :param sheet: 要拆分的工作表
    :param col: 数据列名或列号
    :param begin_row: 数据行开始行号
    :param save_path: 保存路径
    :param keys: 指定只拆分哪些关键字
    :return: None
    """
    Path(save_path).mkdir(parents=True, exist_ok=True)

    if isinstance(keys, str):
        keys = {keys}

    if keys is not None:
        keys = set(keys)

    if isinstance(col, str):
        col = column_index_from_string(col)

    col -= 1

    if begin_row > 1:
        title_rows = sheet[1:begin_row - 1] if begin_row > 2 else (sheet[1],)
    else:
        begin_row = 1
        title_rows = ()

    rows = list(sheet[begin_row:sheet.max_row])
    rows.sort(key=lambda x: str(x[col].value))

    wb = ws = value = None
    for row in rows:
        now_value = '' if row[col].value is None else str(row[col].value)

        if keys is not None and now_value not in keys:
            continue

        # 如果内容和前一个不同，说明前一个已经采集完，可保存并新建工作簿
        if now_value != value:
            if ws is not None:
                wb.save(f'{save_path}\\{make_valid_name(value)}.xlsx')
                wb.close()

            if keys is not None and ws is not None:
                keys.remove(value)
                # 如果keys内没有内容，说明已经采集完毕，退出循环
                if not keys:
                    wb = None
                    break

            wb = Workbook()
            ws = wb.active
            for title in title_rows:
                copy_row(sheet, title, ws)

        copy_row(sheet, row, ws)
        value = now_value

    if wb is not None:
        wb.save(f'{save_path}\\{make_valid_name(value)}.xlsx')
        wb.close()
