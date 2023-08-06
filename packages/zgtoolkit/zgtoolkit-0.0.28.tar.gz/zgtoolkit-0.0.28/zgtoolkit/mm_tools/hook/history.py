from mmcv.runner.hooks import HOOKS, Hook
from visdom import Visdom
import numpy as np
import os.path as osp
import os
from shutil import rmtree
from openpyxl import Workbook, load_workbook
import mmcv
import torch


@HOOKS.register_module()
class History(Hook):
    def __init__(self, metric: list = ['train_loss', 'test_loss', 'lr'], from_scratch=True):
        assert mmcv.is_list_of(metric, str)
        self.viz = Visdom()
        self.from_scratch = from_scratch
        self.metric = metric
        self.iter_train_loss = []
        self.epoch_train_loss = []

        self.iter_val_loss = []
        self.epoch_val_loss = []

    def _save_table(self):
        self.book.save(self.file_path)

    def _save_checkpoint(self, model):
        if self.epoch_val_loss[-1] == min(self.epoch_val_loss):
            torch.save(model.state_dict(), 'train.pth')

    def _insert_table(self, metric):
        # 除了lr，其余metric均保留4位小数
        metric[:-1] = [round(x, 4) for x in metric[:-1]]
        self.sheet.append(metric)
        row = list(self.sheet.rows)[-1]

        for cell in row:
            cell.number_format = 'General'
        self._save_table()

    def before_run(self, runner):
        filename = osp.join(runner.work_dir, 'history')
        self.file_path = osp.join(runner.work_dir, 'history', 'history.xlsx')

        if self.from_scratch:
            self.book = Workbook()
            self.sheet = self.book.active
            self.sheet.append(self.metric)

            if not osp.exists(filename):
                os.mkdir(filename)
            else:
                rmtree(filename)
                os.mkdir(filename)

            self._save_table()
        else:
            self.book = load_workbook(self.file_path)
            self.sheet = self.book.active

            cols = list(self.sheet.iter_cols())
            cols = [[cell.value for cell in col][1:] for col in cols]
            self.viz.line(Y=np.column_stack([np.array(cols[0]), np.array(cols[1])]),
                          X=np.array(range(0, len(cols[0]))),
                          win='line',
                          opts=dict(legend=['train', 'test'])
                          )

            self.epoch_train_loss = cols[0]
            self.epoch_val_loss = cols[1]

    def after_train_iter(self, runner):
        self.iter_train_loss.append(runner.outputs['loss'].item())

    def after_val_iter(self, runner):
        self.iter_val_loss.append(runner.outputs['loss'].item())

    def after_train_epoch(self, runner):
        self.epoch_train_loss.append(np.mean(self.iter_train_loss))
        self.iter_train_loss = []

    def after_val_epoch(self, runner):
        self.epoch_val_loss.append(np.mean(self.iter_val_loss))
        self.iter_val_loss = []

        epoch = len(self.epoch_train_loss)
        lr = runner.optimizer.state_dict()['param_groups'][0]['lr']
        self.viz.line(Y=np.array([[self.epoch_train_loss[-1], self.epoch_val_loss[-1]]]),
                      X=np.array([epoch - 1]),
                      win='line',
                      update='append',
                      opts=dict(legend=['train', 'test']))

        self._insert_table([self.epoch_train_loss[-1],
                            self.epoch_val_loss[-1],
                            lr])

        self._save_checkpoint(runner.model.module)
